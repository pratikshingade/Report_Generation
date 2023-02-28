from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.plotarea import DataTable
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import Reference, Series, BarChart3D

from data_clean_up.clean_data_summary import CleanData


class RawExcel(CleanData):
    def __init__(self, data_path):
        super(RawExcel, self).__init__(data_path)

        self.wb = Workbook()
        self.ws = self.wb.active

    def raw_excel_report(self):
        for r in dataframe_to_rows(self.data_clean(), index=True, header=True):
            self.ws.append(r)

        self.ws["A1"].value = self.ws["A2"].value
        self.ws["A2"].value = ""

    def add_rows(self, row_idx=1, num_of_rows=9):
        self.ws.insert_rows(row_idx, num_of_rows)
        # self.wb.save("test.xlsx")

    def add_chart(self):
        data = Series(Reference(self.ws, min_col=5, min_row=12, max_row=self.ws.max_row - 1),
                      title="ड्रोनद्वारे मोजणी क्षेत्र (Ha)")
        titles = Reference(self.ws, min_col=2, min_row=12, max_row=self.ws.max_row - 1)

        chart = BarChart3D()

        chart.title = "पिक vs क्षेत्र (Ha)"
        chart.y_axis.title = "क्षेत्र (Ha)"
        chart.x_axis.title = 'पिक'

        chart.style = 12
        chart.shape = 'cylinder'
        # chart.add_data(data=data, titles_from_data=True)
        chart.append(data)
        chart.set_categories(titles)
        chart.legend = Legend(legendPos='r')

        chart.plot_area.dTable = DataTable()
        chart.plot_area.dTable.showHorzBorder = True
        chart.plot_area.dTable.showVertBorder = True
        chart.plot_area.dTable.showOutline = True
        chart.plot_area.dTable.showKeys = True

        chart.dLbls = DataLabelList()
        chart.dLblsPos = 'bestfit'

        chart.roundedCorners = True
        chart.height = 10  # default is 7.5
        chart.width = 22

        self.ws.add_chart(chart, "B35")
