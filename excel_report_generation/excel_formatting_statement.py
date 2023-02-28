from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter

from excel_report_generation.raw_excel_statement import RawExcel
from settings import setting_statement


class ExcelFormat(RawExcel):
    def __init__(self, excel_path):
        super(ExcelFormat, self).__init__(excel_path)  # super(Base1, self).__init__()
        self.preferred_title_font = Font(name=setting_statement.FONT, size=setting_statement.TITLE_FONTSIZE,
                                         bold=setting_statement.BOLD_HEADER)
        self.preferred_header_font = Font(name=setting_statement.FONT, size=setting_statement.HEADER_FONTSIZE,
                                          bold=setting_statement.BOLD_HEADER)
        self.preferred_font = Font(name=setting_statement.FONT, size=setting_statement.FONTSIZE,
                                   bold=not setting_statement.BOLD_HEADER)
        self.preferred_border = Border(left=Side(border_style=setting_statement.BORDER_STYLE),
                                       right=Side(border_style=setting_statement.BORDER_STYLE),
                                       top=Side(border_style=setting_statement.BORDER_STYLE),
                                       bottom=Side(border_style=setting_statement.BORDER_STYLE))
        self.preferred_alignment = Alignment(horizontal=setting_statement.CENTER_ALIGNMENT,
                                             vertical=setting_statement.CENTER_ALIGNMENT)

    def printing_setup(self):
        self.ws.print_title_cols = f'{get_column_letter(1)}:{get_column_letter(self.ws.max_column)}'  # the first two
        # cols
        self.ws.print_title_rows = f'{setting_statement.HEADER_ROW_NUMBERS[0]}:{setting_statement.HEADER_ROW_NUMBERS[-1]}'  # the first
        # two rows
        self.ws.print_area = f'{get_column_letter(1)}{setting_statement.HEADER_ROW_NUMBERS[0]}' \
                             f':{get_column_letter(self.ws.max_column)}{self.ws.max_row}'

        self.ws.print_options.horizontalCentered = True
        self.ws.sheet_view.showGridLines = False
        # ws.print_options.verticalCentered = True
        # self.ws.page_setup.fitToPage = True
        # self.ws.page_setup.fitToHeight = True
        # self.ws.page_setup.fitToWidth = True
        self.ws.page_setup.scale = 83 if self.ws.max_column <= 11 else 71 if self.ws.max_column <= 14 else 67

        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE

        self.ws.page_margins.left = setting_statement.PRINT_MARGINS["LEFT"]  # (all margins are in inches)
        self.ws.page_margins.right = setting_statement.PRINT_MARGINS["RIGHT"]

        self.ws.page_margins.top = setting_statement.PRINT_MARGINS["TOP"]
        self.ws.page_margins.bottom = setting_statement.PRINT_MARGINS["BOTTOM"]

        self.ws.page_margins.header = setting_statement.PRINT_MARGINS["HEADER"]
        self.ws.page_margins.footer = setting_statement.PRINT_MARGINS["FOOTER"]
        # return self.ws

    def data_setup(self):
        self.printing_setup()
        self.data_header_cell_formatting()
        self.column_width(setting_statement.COLUMN_WIDTH)
        self.row_height(setting_statement.ROW_HEIGHT)

        crop_statement_row_count = setting_statement.HEADER_ROW_NUMBERS[5]
        while crop_statement_row_count < self.ws.max_row - 1:
            crop_statement_row_count += 1
            data = self.ws.cell(row=crop_statement_row_count, column=setting_statement.COLUMN_NUMBERS[0]).value
            if data:
                for i in range(1, len(self.ws.max_column) + 1):
                    self.data_cell_formatting(crop_statement_row_count, i)
            else:
                break
        self.merging_cell()
        # self.wb.save("test.xlsx")
        return self.ws

    def data_cell_formatting(self, row_number, column_number):
        cell = self.ws.cell(row=row_number, column=column_number)
        cell.font = self.preferred_font
        cell.border = self.preferred_border
        cell.alignment = self.preferred_alignment

    def data_header_cell_formatting(self):
        for cell in self.ws[7] + self.ws[8] + self.ws[self.ws.max_row]:
            cell.font = self.preferred_header_font
            cell.border = self.preferred_border
            cell.alignment = self.preferred_alignment

        for cell_tup in self.ws[9: self.ws.max_row - 1]:
            for cell in cell_tup:
                cell.font = self.preferred_font
                cell.border = self.preferred_border
                cell.alignment = self.preferred_alignment

    def column_width(self, width):
        for col in range(1, self.ws.max_column + 1):
            self.ws.column_dimensions[get_column_letter(col)].width = width

    def row_height(self, height):
        for row in range(1, self.ws.max_row + 1):
            self.ws.row_dimensions[row].height = height

    def merging_cell(self):
        self.ws.merge_cells(f'{get_column_letter(2)}7:{get_column_letter(self.ws.max_column - 2)}7')
        self.ws[
            f'{setting_statement.COLUMN_MAP[2]}{setting_statement.COLUMN_NUMBERS[6]}'].alignment = self.preferred_alignment
        self.ws[
            f'{setting_statement.COLUMN_MAP[2]}{setting_statement.COLUMN_NUMBERS[6]}'].font = self.preferred_header_font
        self.ws[
            f'{setting_statement.COLUMN_MAP[2]}{setting_statement.COLUMN_NUMBERS[6]}'].border = self.preferred_border

        self.ws[f'{setting_statement.COLUMN_MAP[1]}{setting_statement.COLUMN_NUMBERS[6]}'].value = self.ws[
            f'{setting_statement.COLUMN_MAP[1]}'
            f'{setting_statement.COLUMN_NUMBERS[8]}'].value

        self.ws[f'{setting_statement.COLUMN_MAP[1]}{setting_statement.COLUMN_NUMBERS[8]}'].value = ""

        self.ws.merge_cells(
            f"{setting_statement.COLUMN_MAP[1]}{setting_statement.COLUMN_NUMBERS[6]}:{setting_statement.COLUMN_MAP[1]}"
            f"{setting_statement.COLUMN_NUMBERS[7]}")
        self.ws[
            f"{setting_statement.COLUMN_MAP[1]}{setting_statement.COLUMN_NUMBERS[6]}"].alignment = self.preferred_alignment
        self.ws[
            f"{setting_statement.COLUMN_MAP[1]}{setting_statement.COLUMN_NUMBERS[6]}"].border = self.preferred_border
        self.ws[
            f"{setting_statement.COLUMN_MAP[1]}{setting_statement.COLUMN_NUMBERS[6]}"].font = self.preferred_header_font

        self.ws[f"{get_column_letter(self.ws.max_column - 1)}{setting_statement.COLUMN_NUMBERS[6]}"].value = self.ws[
            f"{get_column_letter(self.ws.max_column - 1)}"
            f"{setting_statement.COLUMN_NUMBERS[7]}"].value

        self.ws[f'{get_column_letter(self.ws.max_column - 1)}{setting_statement.COLUMN_NUMBERS[7]}'].value = ""

        self.ws.merge_cells(
            f"{get_column_letter(self.ws.max_column - 1)}{setting_statement.COLUMN_NUMBERS[6]}:{get_column_letter(self.ws.max_column - 1)}"
            f"{setting_statement.COLUMN_NUMBERS[7]}")
        self.ws[
            f"{get_column_letter(self.ws.max_column - 1)}{setting_statement.COLUMN_NUMBERS[6]}"].alignment = self.preferred_alignment
        self.ws[
            f"{get_column_letter(self.ws.max_column - 1)}{setting_statement.COLUMN_NUMBERS[6]}"].border = self.preferred_border
        self.ws[
            f"{get_column_letter(self.ws.max_column - 1)}{setting_statement.COLUMN_NUMBERS[6]}"].font = self.preferred_header_font

        if type(self.ws[f"{get_column_letter(self.ws.max_column)}9"].value) == str:
            self.ws.merge_cells(
                f"{get_column_letter(self.ws.max_column)}{setting_statement.COLUMN_NUMBERS[6]}:{get_column_letter(self.ws.max_column)}"
                f"{setting_statement.COLUMN_NUMBERS[7]}")

            self.ws[
                f'{get_column_letter(self.ws.max_column)}{setting_statement.COLUMN_NUMBERS[6]}'].alignment = self.preferred_alignment
            self.ws[
                f'{get_column_letter(self.ws.max_column)}{setting_statement.COLUMN_NUMBERS[6]}'].font = self.preferred_header_font
            self.ws[
                f'{get_column_letter(self.ws.max_column)}{setting_statement.COLUMN_NUMBERS[6]}'].border = self.preferred_border

            self.ws[f'{get_column_letter(self.ws.max_column)}{setting_statement.COLUMN_NUMBERS[6]}'].value = self.ws[
                f'{get_column_letter(self.ws.max_column)}'
                f'{setting_statement.COLUMN_NUMBERS[8]}'].value

            self.ws[f'{get_column_letter(self.ws.max_column)}{setting_statement.COLUMN_NUMBERS[8]}'].value = ""
