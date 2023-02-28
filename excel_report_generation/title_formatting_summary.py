import os

from openpyxl import load_workbook

from excel_report_generation.excel_formatting_summary import ExcelFormat
from excel_report_generation.raw_excel_summary import RawExcel
from settings import setting_summary


class TitleFormat(ExcelFormat, RawExcel):
    def __init__(self, data_excel_path):
        super(TitleFormat, self).__init__(data_excel_path)
        self.template_path = r"C:\Users\ss\PycharmProjects\Report Generation\excel_report_generation\header " \
                             r"template\summery_template.xlsx"
        if not os.path.exists(self.template_path):
            raise FileNotFoundError("Summary Template File is NOT Found...")

        self.data_excel_path = data_excel_path
        self.template_ws = load_workbook(self.template_path)["Summary Template"]
        # formatted_excel = ExcelFormat(csv_path=self.csv_path)
        self.data_setup()  # formatted_excel.
        self.format_headings()
        self.excel_path = None
        self.data_ws = None

    def format_headings(self):
        for i in range(1, 9):
            for j in range(1, self.ws.max_column + 1):
                c = self.template_ws.cell(row=i, column=j)
                self.ws.cell(row=i, column=j).value = c.value
                self.ws.cell(row=i, column=j).alignment = self.preferred_alignment
                self.ws.cell(row=i, column=j).font = self.preferred_font

        self.ws["C8"].font = self.preferred_title_font
        self.ws.row_dimensions[8].height = setting_summary.TITLE_ROW_HEIGHT
        self.ws["F1"].value += self.format_village_name()
        taluka_name = self.format_taluka_name()
        self.ws["F2"].value += taluka_name
        self.ws["B3"].value += self.format_canal_name()
        self.ws["B2"].value += self.format_office_name()

    def format_village_name(self):
        village_name = os.path.split(os.path.splitext(self.data_excel_path)[0])[1].split("_")[1]
        return village_name

    def format_taluka_name(self):
        tal = os.path.split(os.path.splitext(self.data_excel_path)[0])[1].split("_")[2]
        taluka_name = setting_summary.HEADER_TALUKA_NAME[tal]
        return taluka_name

    def format_canal_name(self):
        canal = os.path.split(os.path.splitext(self.data_excel_path)[0])[1].split("_")[3]
        canal_name = setting_summary.HEADER_CANAL_NAME[canal][0]
        return canal_name

    def format_office_name(self):
        office = os.path.split(os.path.splitext(self.data_excel_path)[0])[1].split("_")[3]
        office_name = setting_summary.HEADER_CANAL_NAME[office][1]
        return office_name


    def save_excel(self):
        village_name_for_file_name = self.format_village_name()
        canal_name_for_file_name = self.format_canal_name()
        excel_path = os.path.split(self.data_excel_path)[0]
        excel_file_name = f"{village_name_for_file_name} - {canal_name_for_file_name} Summary.xlsx"
        self.excel_path = os.path.join(excel_path, excel_file_name)
        self.wb.save(self.excel_path)
