from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter

from data_clean_up.clean_data_statement import CleanData


class RawExcel(CleanData):
    def __init__(self, data_path):
        super(RawExcel, self).__init__(data_path)
        self.data_wb = Workbook()
        self.data_ws = self.data_wb.active

        self.wb = Workbook()
        self.ws = self.wb.active

        # self.raw_excel_report()
        self.source_data_update()
        self.add_rows()

    def raw_excel_report(self):
        data_pivot_table = self.data_clean()
        for r in dataframe_to_rows(data_pivot_table, index=True, header=True):
            self.ws.append(r)

        if self.ws["B3"].value == "स्त्रोत":
            self.ws.move_range(f"B1:B{self.ws.max_row}", cols=self.ws.max_column-1)
            self.ws.delete_cols(2)

        else:
            pass

    def source_data_update(self):
        self.raw_excel_report()

        for row in dataframe_to_rows(self.data, index=False, header=True):
            self.data_ws.append(row)

        for data_cell in self.data_ws[f"A2:{get_column_letter(self.data_ws.max_column)}{self.data_ws.max_row}"]:
            for excel_cell in self.ws[f"A4:{get_column_letter(self.ws.max_column)}{self.ws.max_row}"]:
                if data_cell[0].value == excel_cell[0].value:
                    excel_cell[-1].value = data_cell[3].value

    def add_rows(self, row_idx=1, num_of_rows=6):
        self.ws.insert_rows(row_idx, num_of_rows)
        # self.wb.save("test.xlsx")
