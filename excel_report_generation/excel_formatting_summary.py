from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils.cell import get_column_letter

from excel_report_generation.raw_excel_summary import RawExcel
from settings import setting_summary


class ExcelFormat(RawExcel):
    def __init__(self, excel_path):
        super(ExcelFormat, self).__init__(excel_path)  # super(Base1, self).__init__()
        self.preferred_title_font = Font(name=setting_summary.FONT, size=setting_summary.TITLE_FONTSIZE,
                                         bold=setting_summary.BOLD_HEADER)
        self.preferred_header_font = Font(name=setting_summary.FONT, size=setting_summary.HEADER_FONTSIZE,
                                          bold=setting_summary.BOLD_HEADER)
        self.preferred_font = Font(name=setting_summary.FONT, size=setting_summary.FONTSIZE,
                                   bold=not setting_summary.BOLD_HEADER)
        self.preferred_border = Border(left=Side(border_style=setting_summary.BORDER_STYLE),
                                       right=Side(border_style=setting_summary.BORDER_STYLE),
                                       top=Side(border_style=setting_summary.BORDER_STYLE),
                                       bottom=Side(border_style=setting_summary.BORDER_STYLE))
        self.preferred_alignment = Alignment(horizontal=setting_summary.CENTER_ALIGNMENT,
                                             vertical=setting_summary.CENTER_ALIGNMENT)

    def printing_setup(self):
        self.ws.print_area = f'A1:F58'
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE

        self.ws.print_options.horizontalCentered = True
        self.ws.sheet_view.showGridLines = False
        self.ws.print_options.verticalCentered = True
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToHeight = False
        self.ws.page_setup.fitToWidth = True
        self.ws.page_setup.scale = setting_summary.CUSTOM_PAGE_SCALE

        self.ws.page_margins.left = setting_summary.PRINT_MARGINS["LEFT"]  # (all margins are in inches)
        self.ws.page_margins.right = setting_summary.PRINT_MARGINS["RIGHT"]

        self.ws.page_margins.top = setting_summary.PRINT_MARGINS["TOP"]
        self.ws.page_margins.bottom = setting_summary.PRINT_MARGINS["BOTTOM"]

        self.ws.page_margins.header = setting_summary.PRINT_MARGINS["HEADER"]
        self.ws.page_margins.footer = setting_summary.PRINT_MARGINS["FOOTER"]
        # return self.ws

    def data_setup(self):
        self.raw_excel_report()
        self.printing_setup()
        self.column_width(setting_summary.COLUMN_WIDTH)
        self.row_height(setting_summary.ROW_HEIGHT)
        self.data_header_cell_formatting()
        self.add_rows(1, 9)
        self.set_formula()
        self.add_chart()
        return self.ws

    def set_formula(self):
        self.ws[f"C{self.ws.max_row}"] = f'= SUM(C12:C{self.ws.max_row - 1})'
        self.ws[f"D{self.ws.max_row}"] = f'= SUM(D12:D{self.ws.max_row - 1})'
        self.ws[f"E{self.ws.max_row}"] = f'= SUM(E12:E{self.ws.max_row - 1})'
        self.ws[f"F{self.ws.max_row}"] = f'= SUM(F12:F{self.ws.max_row - 1})'

    def data_header_cell_formatting(self):
        for cell in self.ws[1]:
            cell.font = self.preferred_header_font
            cell.border = self.preferred_border
            cell.alignment = self.preferred_alignment

        for cell_tup in self.ws[f"A2:{get_column_letter(self.ws.max_column)}{self.ws.max_row + 1}"]:
            for _cell in cell_tup:
                _cell.font = self.preferred_font
                _cell.border = self.preferred_border
                _cell.alignment = self.preferred_alignment

        for cell in self.ws[self.ws.max_row]:
            cell.font = self.preferred_header_font
            cell.border = self.preferred_border
            cell.alignment = self.preferred_alignment

    def column_width(self, width):
        for col in range(1, self.ws.max_column + 1):
            self.ws.column_dimensions[get_column_letter(col)].width = width

    def row_height(self, height):
        for row in range(1, self.ws.max_row + 1):
            self.ws.row_dimensions[row].height = height

    def merging_cell(self):
        self.ws.merge_cells(f"A{self.ws.max_row}:B{self.ws.max_row}")
        self.ws[f"A{self.ws.max_row}"].alignment = self.preferred_alignment
        self.ws[f"A{self.ws.max_row}"].font = self.preferred_header_font
        self.ws[f"A{self.ws.max_row}"].border = self.preferred_border
        self.ws[f"A{self.ws.max_row}"].value = "एकूण"
