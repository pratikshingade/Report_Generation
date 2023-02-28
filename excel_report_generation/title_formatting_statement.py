import os

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from excel_report_generation.excel_formatting_statement import ExcelFormat
from excel_report_generation.raw_excel_statement import RawExcel
from settings import setting_statement


class TitleFormat(ExcelFormat, RawExcel):
    def __init__(self, csv_path):
        super(TitleFormat, self).__init__(csv_path)
        self.template_path = r"C:\Users\ss\PycharmProjects\Report Generation\excel_report_generation\header " \
                             r"template\statement_template.xlsx"

        if not os.path.exists(self.template_path):
            raise FileNotFoundError("Statement Template File is NOT Found...")

        self.csv_path = csv_path
        self.template_ws = load_workbook(self.template_path)["Statement Template"]
        # formatted_excel = ExcelFormat(csv_path=self.csv_path)
        self.ws = self.data_setup()  # formatted_excel.
        self.format_headings()
        self.excel_path = None
        self.data_ws = None

    def format_headings(self):
        for i in range(setting_statement.HEADER_ROW_NUMBERS[0], setting_statement.HEADER_ROW_NUMBERS[6]):
            self.ws.row_dimensions[i].height = setting_statement.ROW_HEIGHT
            for j in range(1, self.ws.max_column + 1):
                c = self.template_ws.cell(row=i, column=j)
                self.ws.cell(row=i, column=j).value = c.value
                self.ws.cell(row=i, column=j).alignment = self.preferred_alignment

        self.ws.merge_cells(
            f"{get_column_letter(1)}{setting_statement.HEADER_ROW_NUMBERS[2]}:{get_column_letter(self.ws.max_column)}"
            f"{setting_statement.HEADER_ROW_NUMBERS[2]}")

        self.ws[
            f'{setting_statement.COLUMN_MAP[1]}' 
            f'{setting_statement.HEADER_ROW_NUMBERS[2]}'].value = self.format_canal_name()
        self.ws[
            f'{setting_statement.COLUMN_MAP[1]}' 
            f'{setting_statement.HEADER_ROW_NUMBERS[2]}'].font = self.preferred_title_font
        self.ws.row_dimensions[setting_statement.HEADER_ROW_NUMBERS[2]].height = setting_statement.TITLE_ROW_HEIGHT
        self.ws[
            f'{setting_statement.COLUMN_MAP[1]}' 
            f'{setting_statement.HEADER_ROW_NUMBERS[2]}'].alignment = self.preferred_alignment

        self.ws[
            f'{setting_statement.COLUMN_MAP[6]}' 
            f'{setting_statement.HEADER_ROW_NUMBERS[4]}'].value += self.format_village_name()
        self.ws.column_dimensions[
            f'{setting_statement.COLUMN_MAP[6]}'].width = setting_statement.VILLAGE_NAME_COLUMN_WIDTH

        self.ws[
            f'{setting_statement.COLUMN_MAP[7]}' 
            f'{setting_statement.HEADER_ROW_NUMBERS[4]}'].value += self.format_taluka_name()
        self.ws.column_dimensions[
            f'{setting_statement.COLUMN_MAP[7]}'].width = setting_statement.TALUKA_NAME_COLUMN_WIDTH

        office_name = self.format_office_name()
        self.ws[f"{get_column_letter(9)}1"].value = office_name

    def format_village_name(self):
        village_name = os.path.split(os.path.splitext(self.csv_path)[0])[1].split("_")[1]
        return village_name

    def format_taluka_name(self):
        tal = os.path.split(os.path.splitext(self.csv_path)[0])[1].split("_")[2]
        taluka_name = setting_statement.HEADER_TALUKA_NAME[tal]
        return taluka_name

    def format_canal_name(self):
        canal = os.path.split(os.path.splitext(self.csv_path)[0])[1].split("_")[3]
        canal_name = setting_statement.HEADER_CANAL_NAME[canal][0]
        return canal_name

    def format_office_name(self):
        office = os.path.split(os.path.splitext(self.csv_path)[0])[1].split("_")[3]
        office_name = setting_statement.HEADER_CANAL_NAME[office][1]
        office_name = "उपविभाग : " + office_name
        return office_name

    def save_excel(self):
        village_name_for_file_name = self.format_village_name()
        canal_name_for_file_name = self.format_canal_name()
        excel_path = os.path.split(self.csv_path)[0]
        excel_file_name = f"{village_name_for_file_name} - {canal_name_for_file_name} Statement.xlsx"
        excel_path = os.path.join(excel_path, excel_file_name)
        self.excel_path = excel_path
        self.wb.save(excel_path)
