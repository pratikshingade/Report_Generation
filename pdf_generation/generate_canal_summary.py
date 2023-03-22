import math
import os
import time

import pandas as pd
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart import Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.plotarea import DataTable
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Side, Border, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.pagebreak import RowBreak, Break
from win32com import client

from settings import setting_statement, setting_summary


class GenerateCanalExcel:
    def __init__(self, folder_path):
        self.folder_path = folder_path
        self.root = None
        self.last_file_name = None
        self.last_file_path = None
        self.df = self.read_excel_data()
        self.canal_name_for_file_name = self.format_canal_name()
        self.excel_file_name = f"{self.canal_name_for_file_name} Canal Abstract.xlsx"
        self.excel_path = os.path.join(self.folder_path, self.excel_file_name)
        self.pdf_path = os.path.splitext(self.excel_path)[0] + setting_summary.PDF_FORMAT

        self.template_path = r"C:\Users\ss\PycharmProjects\Report Generation\excel_report" \
                             r"_generation\header template\canal_abstract_template.xlsx"
        if not os.path.exists(self.template_path):
            raise FileNotFoundError("Canal Summary Template File is NOT Found...")

        self.template_ws = load_workbook(self.template_path)["Canal Abstract Template"]

        self.work_book = Workbook()
        self.work_sheet = self.work_book.active

        self.preferred_title_font = Font(name=setting_statement.FONT, size=setting_statement.TITLE_FONTSIZE,
                                         bold=setting_statement.BOLD_HEADER)
        self.preferred_header_font = Font(name=setting_statement.FONT, size=setting_statement.HEADER_FONTSIZE,
                                          bold=setting_statement.BOLD_HEADER)
        self.preferred_font = Font(name=setting_statement.FONT, size=setting_statement.FONTSIZE,
                                   bold=not setting_statement.BOLD_HEADER)
        self.preferred_border = Border(left=Side(border_style=setting_statement.BORDER_STYLE),
                                       right=Side(border_style=setting_statement.BORDER_STYLE),
                                       top=Side(border_style=setting_statement.BORDER_STYLE),
                                       bottom=Side(border_style=setting_statement.BORDER_STYLE))
        self.preferred_alignment = Alignment(horizontal=setting_statement.CENTER_ALIGNMENT,
                                             vertical=setting_statement.CENTER_ALIGNMENT)

    def read_excel_data(self):
        df = pd.DataFrame()
        for root, sub_dirs, files in os.walk(self.folder_path):
            for file in files:
                if file.endswith("Statement.xlsx"):
                    print(os.path.join(root, file))
                    data = pd.read_excel(os.path.join(root, file), header=7)
                    data = data.tail(1)

                    if math.isnan(data.loc[data.index.values[-1], data.columns.values[-1]]):
                        data = data[data.columns[:-1]]

                    data.rename(columns={f"Unnamed: {len(data.axes[1]) - 1}": "मंजुरी क्षेत्र (ha)"}, inplace=True)
                    data["गाव"] = file.split(" ")[0]

                    for col in {"गाव", "Unnamed: 0", "ऊस", "द्राक्षे", "केळी", "गहू", "ज्वारी", "पालेभाजी", "फळभाजी",
                                "बागायत", "मका", "सोयाबीन", "हळद",
                                "इतर", "मंजुरी क्षेत्र (ha)"}.difference(data.columns.values):
                        data[col] = math.nan

                    data = data[
                        ["गाव", "ऊस", "द्राक्षे", "केळी", "गहू", "ज्वारी", "पालेभाजी", "फळभाजी", "बागायत", "मका",
                         "सोयाबीन", "हळद", "इतर", "मंजुरी क्षेत्र (ha)"]]
                    self.last_file_name = file
                    self.root = root
                    self.last_file_path = os.path.join(self.root, self.last_file_name)
                    df = pd.concat([df, data])
        return df

    def raw_excel(self):
        for row in dataframe_to_rows(self.df, index=False, header=True):
            self.work_sheet.append(row)

        return self.work_sheet

    def format_office_name(self):
        wb = load_workbook(self.last_file_path)
        ws = wb.active
        office_name = ws["I1"].value
        office_name = office_name.split(":")[1].strip()
        return office_name

    def format_excel(self):
        for cell in self.work_sheet[1]:
            cell.font = self.preferred_header_font
            cell.border = self.preferred_border
            cell.alignment = self.preferred_alignment

        for cell_tup in self.work_sheet[f"A2:{get_column_letter(self.work_sheet.max_column)}"
                                        f"{self.work_sheet.max_row + 1}"]:
            for _cell in cell_tup:
                _cell.font = self.preferred_font
                _cell.border = self.preferred_border
                _cell.alignment = self.preferred_alignment

        for cell in self.work_sheet[self.work_sheet.max_row]:
            cell.font = self.preferred_header_font
            cell.border = self.preferred_border
            cell.alignment = self.preferred_alignment

        self.row_height(17)
        self.column_width(15)
        self.add_rows(1, 6)
        self.set_formula()
        self.title_formatting()
        self.page_setup()

    def set_formula(self):
        for col in range(2, self.work_sheet.max_column + 1):
            self.work_sheet[f"{get_column_letter(col)}{self.work_sheet.max_row}"] = f'=SUM({get_column_letter(col)}8:' \
                                                                                    f'{get_column_letter(col)}' \
                                                                                    f'{self.work_sheet.max_row - 1})'

        for row in self.work_sheet[f"B9:{get_column_letter(self.work_sheet.max_column)}{self.work_sheet.max_row - 1}"]:
            self.work_sheet[row[-1].coordinate] = f"=SUM({row[0].coordinate}:{row[-2].coordinate})"

    def title_formatting(self):
        for i in range(1, 6):
            for j in range(1, self.work_sheet.max_column + 1):
                c = self.template_ws.cell(row=i, column=j)
                self.work_sheet.cell(row=i, column=j).value = c.value
                self.work_sheet.cell(row=i, column=j).alignment = self.preferred_alignment
                self.work_sheet.cell(row=i, column=j).font = self.preferred_font

        self.work_sheet["J1"].value += self.format_office_name()
        self.work_sheet["A3"].value = self.format_canal_name()
        self.work_sheet.merge_cells(f"A3:{get_column_letter(self.work_sheet.max_column)}3")
        self.work_sheet["A3"].alignment = self.preferred_alignment
        self.work_sheet["A3"].font = self.preferred_title_font
        self.work_sheet.row_dimensions[3].height = 29

    def format_canal_name(self):
        canal_name = os.path.splitext(self.last_file_name)[0].split(" - ")[1].split(" Statement")[0]
        return canal_name

    def column_width(self, width):
        for col in range(1, self.work_sheet.max_column + 1):
            self.work_sheet.column_dimensions[get_column_letter(col)].width = width

    def row_height(self, height):
        for row in range(1, self.work_sheet.max_row + 1):
            self.work_sheet.row_dimensions[row].height = height

    def add_rows(self, row_idx=1, num_of_rows=6):
        self.work_sheet.insert_rows(row_idx, num_of_rows)

    def save_excel(self):
        self.raw_excel()
        self.format_excel()
        self.add_chart()
        self.work_book.save(self.excel_path)

    def page_setup(self):
        self.work_sheet.page_setup.paperSize = self.work_sheet.PAPERSIZE_A4
        self.work_sheet.page_setup.orientation = self.work_sheet.ORIENTATION_LANDSCAPE
        self.work_sheet.print_area = f'A1:N58'

        self.work_sheet.print_options.horizontalCentered = True
        self.work_sheet.sheet_view.showGridLines = False
        self.work_sheet.print_options.verticalCentered = True
        self.work_sheet.page_setup.fitToPage = True
        self.work_sheet.page_setup.fitToHeight = False
        self.work_sheet.page_setup.fitToWidth = True

        row_number = 33  # the row that you want to insert page break
        row_break = RowBreak()
        row_break.append(Break(id=row_number))
        self.work_sheet.row_breaks = row_break  # insert page break
        self.work_sheet.page_setup.scale = setting_summary.CUSTOM_PAGE_SCALE

        self.work_sheet.page_margins.left = setting_summary.PRINT_MARGINS["LEFT"]  # (all margins are in inches)
        self.work_sheet.page_margins.right = setting_summary.PRINT_MARGINS["RIGHT"]

        self.work_sheet.page_margins.top = setting_summary.PRINT_MARGINS["TOP"]
        self.work_sheet.page_margins.bottom = setting_summary.PRINT_MARGINS["BOTTOM"]

        self.work_sheet.page_margins.header = setting_summary.PRINT_MARGINS["HEADER"]
        self.work_sheet.page_margins.footer = setting_summary.PRINT_MARGINS["FOOTER"]

    def add_chart(self):
        measurement = []
        crop = []
        for i in range(2, self.work_sheet.max_column):
            value = self.work_sheet.cell(row=self.work_sheet.max_row, column=i).value
            measurement.append(value)

            value = self.work_sheet.cell(row=7, column=i).value
            crop.append(value)

        for j, v, k in zip(range(35, 35 + self.work_sheet.max_column), measurement, crop):
            # print(j, v, k)
            self.work_sheet.cell(row=j, column=3).value = v
            self.work_sheet.cell(row=j, column=2).value = k

        data = Series(Reference(self.work_sheet, min_col=3, max_col=3, min_row=35, max_row=self.work_sheet.max_row),
                      title="ड्रोनद्वारे मोजणी क्षेत्र (Ha)")
        titles = Reference(self.work_sheet, min_col=2, max_col=2, min_row=35, max_row=self.work_sheet.max_row)

        chart = BarChart3D()

        chart.title = f"{self.canal_name_for_file_name} पिकांची वर्गवारी"
        chart.y_axis.title = "क्षेत्र (Ha)"
        chart.x_axis.title = 'पिक'

        chart.shape = 'cylinder'
        # chart.add_data(data=data, titles_from_data=True)
        chart.append(data)
        chart.set_categories(titles)
        chart.legend = Legend(legendPos='r')

        chart.plot_area.dTable = DataTable()
        chart.plot_area.dTable.showHorzBorder = True
        chart.plot_area.dTable.showVertBorder = True
        chart.plot_area.dTable.showOutline = True
        chart.plot_area.dTable.showKeys = True

        chart.dLbls = DataLabelList()
        chart.dLblsPos = 'bestfit'

        chart.roundedCorners = True
        chart.height = 12  # default is 7.5
        chart.width = 30

        self.work_sheet.add_chart(chart, "B35")

    def create_pdf(self):
        self.save_excel()
        excel = client.Dispatch(setting_summary.PDF_GENERATION_APPLICATION)
        excel.Interactive = False
        excel.Visible = False
        wb = excel.Workbooks.Open(self.excel_path)
        ws = wb.Worksheets[0]

        print(self.pdf_path)
        ws.ExportAsFixedFormat(0, self.pdf_path)
        wb.Close(True)
        time.sleep(1.5)
